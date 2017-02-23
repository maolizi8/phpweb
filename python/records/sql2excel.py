# -*- coding: utf-8 -*-
import pymysql as mc  
from openpyxl.writer.excel import ExcelWriter  
from openpyxl.workbook import Workbook  
#it's different between workbook and worksheet  
conn = mc.connect(user='root', password='', host='127.0.0.1', database='maolizi',charset="utf8")  
cur = conn.cursor()  
  
# query language  
#sql = "select name,age from user" 
sql ="""SELECT issues.issuenumber as '关键字',issues.summary as '摘要',issues.priority as '优先级',issues.status as '状态',issues.fix as '解决',issues.created as '创建日期',userone.assign as '经办人',usertwo.report as '报告人' FROM
(
(
SELECT j.id,j.PROJECT,CONCAT(p.pkey,'-',j.issuenum) as issuenumber,j.SUMMARY,j.ASSIGNEE,j.REPORTER,pi.pname as priority,i.pname as status,r.pname as fix,j.CREATED
from jiraissue j,project p,resolution r,issuestatus i,priority pi
where j.PROJECT=p.id
AND j.PRIORITY=pi.ID
AND j.issuestatus=i.ID
AND j.RESOLUTION=r.ID
AND  j.issuetype=1
AND (j.PRIORITY=1 OR j.PRIORITY=2)
AND j.project=12503
) issues
LEFT JOIN
(
SELECT a.user_key,c.display_name as assign from app_user a,cwd_user c
WHERE a.lower_user_name=c.lower_user_name
) userone on issues.ASSIGNEE=userone.user_key
LEFT JOIN
(
SELECT a.user_key,c.display_name as report from app_user a,cwd_user c
WHERE a.lower_user_name=c.lower_user_name
) usertwo on issues.REPORTER=usertwo.user_key
)"""   
data = cur.execute(sql)  
index = cur.description 

# use fetchall method to get a list, some tuples in it  
data_list = cur.fetchall() 
cur.close()  
conn.close()
# print (data_list)  
  
wb = Workbook()  
ws = wb.worksheets[0]  
ws.title = 'data'  
  
file_name = r'user.xlsx'  
rows = len(data_list)  
cols = len(data_list[0])  
  
# same as read data  
for rx in range(rows):  
	for cx in range(cols): 
		if rx==0:
			ws.cell(row=1,column=cx+1).value=index[cx][0]
		ws.cell(row=rx+2, column=cx+1).value = data_list[rx][cx]  
  
# use ExcelWriter method to write file and save it  
wb.save('user.xlsx') 
